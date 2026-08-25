"""
import_daily_submissions.py - Reads a daily submissions Excel file and adds
new clients to the "sales" table (PostgreSQL / Supabase version), skipping
any already-imported client (matched by National ID + Departure Date).
"""
import os
import re
from datetime import date, datetime
from openpyxl import load_workbook

from db import get_connection
from pricing_engine import calculate_row, build_pricing_cache

REQUIRED_HEADERS = ["الاسم", "تاريخ الميلاد", "الرقم القومي", "رقم الجواز", "المنفذ",
                    "جهه المغادره", "رقم الرحله", "تاريخ المغادرة", "الوكيل",
                    "ملاحظات", "نوع الخدمه"]


def to_iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date().isoformat()
    except ValueError:
        return None


def extract_date_from_filename(fname):
    m = re.search(r"من.*?(\d{4}-\d{2}-\d{2})", fname)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d").date().isoformat()
        except ValueError:
            pass
    return date.today().isoformat()


def import_file(xlsx_path, performed_by=""):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            h = str(h).strip()
            if h and h not in headers:
                headers[h] = c

    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        return {"success": False, "error": f"الملف ناقص الأعمدة التالية: {', '.join(missing)}"}

    submit_date = extract_date_from_filename(os.path.basename(xlsx_path))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT national_id, departure_date FROM sales")
    existing_keys = set(cur.fetchall())

    pricing_cache = build_pricing_cache(cur)
    new_price_placeholders = set()

    # ---- تحميل الوكلاء/شركات الطيران/موردى الاستثمار الموجودين فعلاً فى الذاكرة
    # مرة واحدة بس - بدل ما نسأل قاعدة البيانات لكل صف لو الاسم موجود قبل كده،
    # وده كان السبب الرئيسى فى بطء الاستيراد وانتهاء وقت الطلب (timeout) على
    # الاستضافة، خصوصًا مع ملفات فيها مئات العملاء ونفس الوكيل/الشركة متكررين ----
    cur.execute("SELECT name FROM agents")
    known_agents = {r[0] for r in cur.fetchall()}
    cur.execute("SELECT name FROM airlines")
    known_airlines = {r[0] for r in cur.fetchall()}
    cur.execute("SELECT name FROM investment_suppliers")
    known_investment_suppliers = {r[0] for r in cur.fetchall()}

    added = 0
    skipped = 0
    package_counts = {}
    zero_price_count = 0
    total_sales_value = 0.0
    inserted_ids = []

    r = 2
    while ws.cell(row=r, column=headers["الاسم"]).value is not None:
        nid = str(ws.cell(row=r, column=headers["الرقم القومي"]).value or "").strip()
        if nid == "":
            r += 1
            continue

        dep_date = to_iso(ws.cell(row=r, column=headers["تاريخ المغادرة"]).value)
        key = (nid, dep_date)
        if key in existing_keys:
            skipped += 1
            r += 1
            continue

        name = str(ws.cell(row=r, column=headers["الاسم"]).value or "").strip()
        dob = to_iso(ws.cell(row=r, column=headers["تاريخ الميلاد"]).value)
        passport = str(ws.cell(row=r, column=headers["رقم الجواز"]).value or "").strip()
        port = str(ws.cell(row=r, column=headers["المنفذ"]).value or "").strip()
        dest = str(ws.cell(row=r, column=headers["جهه المغادره"]).value or "").strip()
        flight = str(ws.cell(row=r, column=headers["رقم الرحله"]).value or "").strip()
        agent = str(ws.cell(row=r, column=headers["الوكيل"]).value or "").strip()
        notes = str(ws.cell(row=r, column=headers["ملاحظات"]).value or "").strip()
        package = str(ws.cell(row=r, column=headers["نوع الخدمه"]).value or "").strip()

        category = notes if notes in ("طفل", "رضيع", "انثى") else "بالغ"

        # ---- لو التركيبة (باكدج+خط سير+فئة) دي جديدة تمامًا (مفيش أي صف ليها
        # فى جدول أسعار البيع)، نضيفها تلقائيًا بسعر صفر - تظهر جاهزة فى شاشة
        # "تعديل أسعار البيع" عشان تُكتب فيها السعر الحقيقى بدل ما تختفي بصمت ----
        price_key = (package, port, dest, category)
        if price_key not in pricing_cache["sell_prices"] and all(price_key):
            cur.execute("""
                INSERT INTO total_sell_prices(date_from, date_to, package_code, port, destination, category, total_price)
                VALUES ('2020-01-01', '2099-12-31', %s, %s, %s, %s, 0)
            """, (package, port, dest, category))
            pricing_cache["sell_prices"][price_key] = [("2020-01-01", "2099-12-31", 0.0)]
            new_price_placeholders.add(price_key)

        row_data = calculate_row(
            cur, name, dob, nid, passport, port, dest, flight, dep_date,
            submit_date, agent, category, package, cache=pricing_cache
        )

        if agent and agent not in known_agents:
            cur.execute("INSERT INTO agents(name) VALUES (%s) ON CONFLICT (name) DO NOTHING", (agent,))
            known_agents.add(agent)
        if row_data["airline"] and row_data["airline"] not in known_airlines:
            cur.execute("INSERT INTO airlines(name) VALUES (%s) ON CONFLICT (name) DO NOTHING",
                        (row_data["airline"],))
            known_airlines.add(row_data["airline"])
        if row_data["investment_supplier"] and row_data["investment_supplier"] not in known_investment_suppliers:
            cur.execute("INSERT INTO investment_suppliers(name) VALUES (%s) ON CONFLICT (name) DO NOTHING",
                        (row_data["investment_supplier"],))
            known_investment_suppliers.add(row_data["investment_supplier"])

        cur.execute("""
            INSERT INTO sales(
                name, date_of_birth, national_id, passport_number, port, destination,
                flight_number, departure_date, submission_date, agent, investment_supplier,
                category, package_code, airline,
                service_price, ticket_price, total_sales,
                visa_cost, investment_cost, approval_cost, service_cost_total, ticket_cost, total_cost, net_profit,
                booking_status, no_show_penalty
            ) VALUES (%(name)s,%(date_of_birth)s,%(national_id)s,%(passport_number)s,%(port)s,%(destination)s,
                      %(flight_number)s,%(departure_date)s,%(submission_date)s,%(agent)s,%(investment_supplier)s,
                      %(category)s,%(package_code)s,%(airline)s,
                      %(service_price)s,%(ticket_price)s,%(total_sales)s,
                      %(visa_cost)s,%(investment_cost)s,%(approval_cost)s,%(service_cost_total)s,%(ticket_cost)s,%(total_cost)s,%(net_profit)s,
                      %(booking_status)s,%(no_show_penalty)s)
            RETURNING id
        """, row_data)
        new_id = cur.fetchone()[0]

        existing_keys.add(key)
        inserted_ids.append(new_id)
        added += 1
        package_counts[package] = package_counts.get(package, 0) + 1
        if row_data["total_sales"] <= 0:
            zero_price_count += 1
        total_sales_value += row_data["total_sales"]
        r += 1

    min_id = min(inserted_ids) if inserted_ids else None
    max_id = max(inserted_ids) if inserted_ids else None
    cur.execute("""
        INSERT INTO import_log(source_file, rows_added, rows_skipped, performed_by, min_id, max_id)
        VALUES (%s,%s,%s,%s,%s,%s)
    """, (os.path.basename(xlsx_path), added, skipped, performed_by, min_id, max_id))

    conn.commit()
    conn.close()

    return {
        "success": True, "added": added, "skipped": skipped, "submit_date": submit_date,
        "package_counts": package_counts, "zero_price_count": zero_price_count,
        "total_sales_value": total_sales_value,
        "new_price_combos": [
            {"package": p, "port": port, "destination": d, "category": c}
            for (p, port, d, c) in new_price_placeholders
        ],
    }
