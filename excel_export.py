"""
excel_export.py - Generates downloadable Excel files (in-memory, no disk
writes needed since Render's filesystem is ephemeral) for the web app's
reports: full report, account statement, treasury report, sales by period.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import get_connection
import reports

BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
GREEN_FILL = PatternFill('solid', fgColor='C6E0B4')
BLACK = Font(name='Arial', color='000000')
BLACK_BOLD = Font(name='Arial', bold=True, color='000000')
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


def _new_sheet(wb, title, tab_color=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    return ws


def _write_title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    ws.row_dimensions[1].height = 28


def _write_table(ws, start_row, headers, rows, col_widths, total_cols=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    r = start_row + 1
    for row in rows:
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if i > 1 else RIGHT
            cell.font = BLACK
            if isinstance(val, float):
                cell.number_format = "#,##0"
        r += 1
    if total_cols and rows:
        ws.cell(row=r, column=1, value="الإجمالى").font = BLACK_BOLD
        ws.cell(row=r, column=1).border = BORDER
        for col_idx in total_cols:
            letter = get_column_letter(col_idx)
            tcell = ws.cell(row=r, column=col_idx, value=f"=SUM({letter}{start_row+1}:{letter}{r-1})")
            tcell.font = BLACK_BOLD
            tcell.fill = GREEN_FILL
            tcell.border = BORDER
            tcell.number_format = "#,##0"
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return r + 2


def build_full_report_xlsx(date_from="2020-01-01", date_to="2099-12-31"):
    conn = get_connection()
    cur = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    ws = _new_sheet(wb, "حساب الوكلاء", "2E7D32")
    _write_title(ws, "حساب العملاء (الوكلاء)", 4)
    agent_results = reports.get_all_agent_accounts(date_from, date_to)
    rows = [(r["agent_name"], r["total_sales"], r["total_collected"], r["balance"]) for r in agent_results]
    _write_table(ws, 3, ["اسم الوكيل", "إجمالى المبيعات", "إجمالى المحصَّل", "الرصيد المتبقى"],
                 rows, [26, 18, 18, 18], total_cols=[2, 3, 4])

    ws = _new_sheet(wb, "حساب موردى التأشيرات", "6A1B9A")
    _write_title(ws, "حساب موردى التأشيرات", 4)
    cur.execute("SELECT name FROM visa_suppliers ORDER BY name")
    rows = [(n, (a := reports.get_visa_supplier_account(cur, n, date_from, date_to))["total_cost"],
              a["total_paid"], a["balance"]) for (n,) in cur.fetchall()]
    _write_table(ws, 3, ["اسم المورد", "إجمالى التكلفة", "إجمالى المسدَّد", "الرصيد المتبقى"],
                 rows, [26, 18, 18, 18], total_cols=[2, 3, 4])

    ws = _new_sheet(wb, "حساب موردى الاستثمار", "AD1457")
    _write_title(ws, "حساب موردى الاستثمار", 4)
    cur.execute("SELECT name FROM investment_suppliers ORDER BY name")
    rows = [(n, (a := reports.get_investment_supplier_account(cur, n, date_from, date_to))["total_cost"],
              a["total_paid"], a["balance"]) for (n,) in cur.fetchall()]
    _write_table(ws, 3, ["اسم المورد", "إجمالى التكلفة", "إجمالى المسدَّد", "الرصيد المتبقى"],
                 rows, [26, 18, 18, 18], total_cols=[2, 3, 4])

    ws = _new_sheet(wb, "حساب شركات الطيران", "F57F17")
    _write_title(ws, "حساب شركات الطيران", 5)
    cur.execute("SELECT name FROM airlines ORDER BY name")
    rows = [(n, (a := reports.get_airline_account(cur, n, date_from, date_to))["ticket_count"],
              a["total_cost"], a["total_paid"], a["balance"]) for (n,) in cur.fetchall()]
    _write_table(ws, 3, ["شركة الطيران", "عدد التذاكر", "إجمالى التكلفة", "إجمالى المسدَّد", "الرصيد المتبقى"],
                 rows, [22, 14, 18, 18, 18], total_cols=[3, 4, 5])

    ws = _new_sheet(wb, "حساب الشركاء", "00838F")
    _write_title(ws, "حساب الشركاء", 5)
    partner_results = reports.get_all_partner_accounts(date_from, date_to)
    rows = [(r["partner_name"], f"{r['share_percentage']*100:.1f}%", r["entitled"],
             r["distributed"], r["remaining"]) for r in partner_results]
    _write_table(ws, 3, ["اسم الشريك", "النسبة", "نصيبه من الربح", "الموزَّع له", "المتبقى له"],
                 rows, [22, 12, 18, 18, 18], total_cols=[3, 4, 5])

    ws = _new_sheet(wb, "الأرباح والخسائر", "B71C1C")
    _write_title(ws, "ملخص الأرباح والخسائر", 2)
    pl = reports.get_profit_and_loss(cur, date_from, date_to)
    pl_rows = [
        ("إيرادات الفترة", pl["period_revenue"]), ("إيرادات الأرصدة الافتتاحية", pl["opening_revenue"]),
        ("إجمالى الإيرادات", pl["total_revenue"]), ("", ""),
        ("تكلفة الفترة", pl["period_cost"]), ("تكلفة الأرصدة الافتتاحية", pl["opening_cost"]),
        ("إجمالى التكلفة", pl["total_cost"]), ("", ""),
        ("المصروفات العمومية", pl["general_expenses"]), ("", ""),
        ("صافى الربح", pl["net_profit"]),
    ]
    r = 3
    for label, val in pl_rows:
        is_bold = label in ("إجمالى الإيرادات", "إجمالى التكلفة", "صافى الربح")
        font = BLACK_BOLD if is_bold else BLACK
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font
        lc.alignment = RIGHT
        lc.border = BORDER
        if val != "":
            vc = ws.cell(row=r, column=2, value=val)
            vc.font = font
            vc.number_format = "#,##0"
            vc.alignment = CENTER
            vc.border = BORDER
        r += 1
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def clean_filename(s):
    for ch in '\\/:*?"<>|':
        s = s.replace(ch, "_")
    return s.strip()


def build_statement_xlsx(entity_name, entity_type, summary, txn_headers, txn_rows, all_time):
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف حساب"
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [14, 26, 16, 18, 22, 16]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"كشف حساب: {entity_name}")
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    ws.row_dimensions[1].height = 28
    ws.cell(row=2, column=1, value=f"تاريخ الإصدار: {datetime.now().strftime('%Y-%m-%d')}").font = BLACK

    r = 4
    ws.cell(row=r, column=1, value="أولاً: الملخص").font = BLACK_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for key, val in summary.items():
        ws.cell(row=r, column=1, value=key).font = BLACK
        ws.cell(row=r, column=1).alignment = RIGHT
        ws.cell(row=r, column=1).border = BORDER
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = BLACK_BOLD
        vc.border = BORDER
        vc.alignment = CENTER
        if isinstance(val, float):
            vc.number_format = "#,##0"
        r += 1

    if txn_headers:
        r += 1
        ws.cell(row=r, column=1, value="ثانياً: تفاصيل المعاملات").font = BLACK_BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        for i, h in enumerate(txn_headers, start=1):
            hc = ws.cell(row=r, column=i, value=h)
            hc.fill = HEADER_FILL
            hc.font = HEADER_FONT
            hc.alignment = CENTER
            hc.border = BORDER
        r += 1
        for row in txn_rows:
            for i, val in enumerate(row.values(), start=1):
                cell = ws.cell(row=r, column=i, value=val)
                cell.border = BORDER
                cell.alignment = CENTER
                if isinstance(val, float):
                    cell.number_format = "#,##0"
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="ثالثاً: الإجمالى الكلي (كل الفترات)").font = BLACK_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for key, val in all_time.items():
        ws.cell(row=r, column=1, value=key).font = BLACK
        ws.cell(row=r, column=1).alignment = RIGHT
        ws.cell(row=r, column=1).border = BORDER
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = BLACK_BOLD
        vc.border = BORDER
        vc.alignment = CENTER
        if isinstance(val, float):
            vc.number_format = "#,##0"
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
